import os
import openpyxl
from random import choice
from xmlrpc.server import SimpleXMLRPCServer
import time

primary_metadata = {}
servers_metadata = {}
backup_servers = {}
masterServer = SimpleXMLRPCServer(('localhost', 9000), logRequests=True, allow_none=True)

def send_backup_servers(backup):
    for server in backup:
        filename = server[0]
        addr = server[1]
        port = server[2]
        timestamp = server[3]

        if filename in backup_servers:
            backup_servers[filename].add((addr, port, timestamp))
        else:
            backup_servers[filename] = {(addr, port, timestamp)}

    print(backup_servers)

def lock(filename):
    if filename in primary_metadata:
        if primary_metadata[filename]["status"] == "unlocked":
            primary_metadata[filename]["status"] = "locked"
            return True, primary_metadata[filename]["addr"], primary_metadata[filename]["port"], time.time()
        else:
            return False, None, None, None
    else:
        server_data = choice(list(servers_metadata.values()))
        primary_metadata[filename] = {
            "id": generate_id(),
            "addr": server_data["addr"],
            "port": server_data["port"],
            "status": "locked"
        }
        return True, server_data["addr"], server_data["port"], time.time()

def unlock(filename):
    if filename in primary_metadata and primary_metadata[filename]["status"] == "locked":
        primary_metadata[filename]["status"] = "unlocked"
        return True
    else:
        return False

def write(filename):
    success, addr, port, timestamp = lock(filename)
    if success:
        return True, addr, port, timestamp
    else:
        return False, None, None, None

def read(filename):
    read_servers = []
    if filename in backup_servers:
        for addr, port, timestamp in backup_servers[filename]:
            read_servers.append((addr, port))
    return read_servers

def generate_id():
    return len(primary_metadata) + 1

def load_servers():
    server_file = "Servers.xlsx"
    if os.path.exists(server_file):
        workbook = openpyxl.load_workbook(server_file)
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            name, addr, port = row
            servers_metadata[name] = {"addr": addr, "port": port}
    else:
        print("Servers.xlsx file not found.")

def load_primary_metadata():
    primary_metadata_file = "Primary_Metadata.xlsx"
    if os.path.exists(primary_metadata_file):
        workbook = openpyxl.load_workbook(primary_metadata_file)
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            filename, addr, port, status = row
            primary_metadata[filename] = {"addr": addr, "port": port, "status": status}

load_servers()
load_primary_metadata()

masterServer.register_function(write, 'write')
masterServer.register_function(read, 'read')
masterServer.register_function(lock, 'lock')
masterServer.register_function(unlock, 'unlock')
masterServer.register_function(send_backup_servers, 'send_backup_servers')

print("Master server is running...")
masterServer.serve_forever()
